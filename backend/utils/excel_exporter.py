from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
from backend.utils.logger import logger

OUTPUT_DIR = "outputs"


def export_to_excel(items: list[dict], product_name: str) -> str | None:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = product_name.replace("/", "_").replace("\\", "_")
    excel_path = os.path.join(OUTPUT_DIR, f"{date_str}_{safe_name}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "AI Product Content"
    ws.append(["Item ID", "Type", "Content", "Status"])

    for item in items:
        ws.append([
            item.get("item_id", ""),
            item.get("item_type", item.get("type", "")),
            item.get("content", ""),
            item.get("status", "Pending"),
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 60)

    try:
        wb.save(excel_path)
        logger.info(f"Excel导出成功: {excel_path}")
        return excel_path
    except PermissionError:
        logger.error("Excel导出失败：文件被占用")
        return None


def update_excel_status(excel_path: str, item_id: str, new_status: str) -> bool:
    if not os.path.exists(excel_path):
        return False
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == item_id:
                row[3].value = new_status
                wb.save(excel_path)
                return True
        return False
    except PermissionError:
        logger.error("Excel更新失败：文件被占用")
        return False
