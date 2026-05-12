from openpyxl import Workbook, load_workbook
import os
from datetime import datetime


OUTPUT_DIR = "outputs"


def create_excel_filename(product_name):
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = product_name.replace("/", "_").replace("\\", "_")
    return f"{OUTPUT_DIR}/{date_str}_{safe_name}.xlsx"


def export_to_excel(items, product_name):
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    excel_path = create_excel_filename(product_name)

    wb = Workbook()
    ws = wb.active
    ws.title = "AI Product Content"

    ws.append(["Item ID", "Type", "Content", "Status"])

    for item in items:
        ws.append([
            item["item_id"],
            item["type"],
            item["content"],
            item["status"]
        ])

    try:
        wb.save(excel_path)
        return excel_path
    except PermissionError:
        return None


def update_excel_status(excel_path, item_id, new_status):
    if not os.path.exists(excel_path):
        return False

    wb = load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        excel_item_id = row[0].value

        if excel_item_id == item_id:
            row[3].value = new_status

            try:
                wb.save(excel_path)
                return True
            except PermissionError:
                return False

    return False